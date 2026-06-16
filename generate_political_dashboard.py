import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_political_dashboard():
    filename = "Political_Parties_Master_Dashboard.xlsx"
    
    # 1. Clean, structured datasets directly parsed from the copied site dumps
    summary_data = [
        {
            "Party": "Congress (INC)",
            "IG Followers": 13952620,
            "IG Media Count": 40610,
            "IG Engagement Rate": 0.0014, # Convert 0.14% to formal fraction
            "IG Avg Likes": 19507.50,
            "IG Avg Comments": 280.94,
            "FB Total Likes": 11377074,
            "FB Talking About": 5909662,
            "YT Subscribers": 7360000,
            "YT Video Count": 66129,
            "YT Total Views": 4627576769
        },
        {
            "Party": "Bharatiya Janata Party (BJP)",
            "IG Followers": 9464567,
            "IG Media Count": 18911,
            "IG Engagement Rate": 0.0006, # Convert 0.06% to formal fraction
            "IG Avg Likes": 5271.56,
            "IG Avg Comments": 74.69,
            "FB Total Likes": 18303467,
            "FB Talking About": 3435598,
            "YT Subscribers": 6380000,
            "YT Video Count": 56059,
            "YT Total Views": 2782959031
        },
        {
            "Party": "Aam Aadmi Party (AAP)",
            "IG Followers": 1959577,
            "IG Media Count": 13618,
            "IG Engagement Rate": 0.0016, # Convert 0.16% to formal fraction
            "IG Avg Likes": 2953.88,
            "IG Avg Comments": 96.00,
            "FB Total Likes": 6221880,
            "FB Talking About": 1090208,
            "YT Subscribers": 7520000,
            "YT Video Count": 21605,
            "YT Total Views": 4481437405
        }
    ]
    df_summary = pd.DataFrame(summary_data)
    
    # Platform-specific recent 14-day tracking velocity metrics 
    growth_data = [
        {"Party": "INC", "Platform": "Instagram", "14D Net Growth": 258299, "14D Content Output": 430},
        {"Party": "BJP", "Platform": "Instagram", "14D Net Growth": 37198, "14D Content Output": 269},
        {"Party": "AAP", "Platform": "Instagram", "14D Net Growth": 9376, "14D Content Output": 113},
        
        {"Party": "INC", "Platform": "Facebook", "14D Net Growth": 193576, "14D Content Output": "N/A"},
        {"Party": "BJP", "Platform": "Facebook", "14D Net Growth": 70358, "14D Content Output": "N/A"},
        {"Party": "AAP", "Platform": "Facebook", "14D Net Growth": 107961, "14D Content Output": "N/A"},
        
        {"Party": "INC", "Platform": "YouTube", "14D Net Growth": 60000, "14D Content Output": 783},
        {"Party": "BJP", "Platform": "YouTube", "14D Net Growth": 0, "14D Content Output": 339},
        {"Party": "AAP", "Platform": "YouTube", "14D Net Growth": 10000, "14D Content Output": 155}
    ]
    df_growth = pd.DataFrame(growth_data)
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        pd.DataFrame().to_excel(writer, sheet_name="Master Dashboard")
        workbook = writer.book
        sheet = workbook["Master Dashboard"]
        sheet.views.sheetView[0].showGridLines = True
        
        # UI Styling Tokens
        INDIGO_DARK = "1A2B4C"
        INDIGO_LIGHT = "F4F6F9"
        SAFFRON = "FF9933"
        GREEN = "138808"
        WHITE = "FFFFFF"
        TEXT_MUTED = "555555"
        BORDER_GRAY = "D1D5DB"
        
        font_title = Font(name="Segoe UI", size=18, bold=True, color=INDIGO_DARK)
        font_subtitle = Font(name="Segoe UI", size=10, italic=True, color=TEXT_MUTED)
        font_section = Font(name="Segoe UI", size=13, bold=True, color=INDIGO_DARK)
        font_header = Font(name="Segoe UI", size=11, bold=True, color=WHITE)
        font_data = Font(name="Segoe UI", size=11, color="333333")
        
        fill_indigo = PatternFill(start_color=INDIGO_DARK, end_color=INDIGO_DARK, fill_type="solid")
        fill_zebra = PatternFill(start_color=INDIGO_LIGHT, end_color=INDIGO_LIGHT, fill_type="solid")
        fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color=BORDER_GRAY), right=Side(style='thin', color=BORDER_GRAY),
            top=Side(style='thin', color=BORDER_GRAY), bottom=Side(style='thin', color=BORDER_GRAY)
        )
        border_saffron_top = Border(top=Side(style='medium', color=SAFFRON), bottom=Side(style='thin', color=BORDER_GRAY), left=Side(style='thin', color=BORDER_GRAY), right=Side(style='thin', color=BORDER_GRAY))
        border_green_bottom = Border(bottom=Side(style='medium', color=GREEN), top=Side(style='thin', color=BORDER_GRAY), left=Side(style='thin', color=BORDER_GRAY), right=Side(style='thin', color=BORDER_GRAY))
        
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        
        # --- TITLE BANNER ---
        sheet["B2"] = "NATIONAL POLITICAL PARTIES SOCIAL MEDIA AUDIT INDEX"
        sheet["B2"].font = font_title
        sheet["B3"] = "Cross-Platform Performance Analysis Grid • Data Verified June 2026"
        sheet["B3"].font = font_subtitle
        sheet.row_dimensions[2].height = 26
        sheet.row_dimensions[3].height = 15
        
        # --- SECTION I: GLOBAL COMPARISON MATRIX ---
        sheet["B5"] = "I. Cross-Party Global Performance Matrix"
        sheet["B5"].font = font_section
        
        headers_sum = list(df_summary.columns)
        for col_idx, header in enumerate(headers_sum, start=2):
            cell = sheet.cell(row=6, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_indigo
            cell.alignment = align_center
            cell.border = border_saffron_top
        sheet.row_dimensions[6].height = 26
        
        current_row = 7
        for row_idx, row_data in enumerate(df_summary.values):
            row_fill = fill_white if row_idx % 2 == 0 else fill_zebra
            for c_idx, value in enumerate(row_data, start=2):
                cell = sheet.cell(row=current_row, column=c_idx, value=value)
                cell.font = font_data
                cell.border = thin_border
                cell.fill = row_fill
                
                if c_idx == 2:
                    cell.alignment = align_left
                elif c_idx == 5:
                    cell.number_format = '0.00%'
                    cell.alignment = align_right
                elif c_idx in [6, 7]:
                    cell.number_format = '#,##0.00'
                    cell.alignment = align_right
                else:
                    cell.number_format = '#,##0'
                    cell.alignment = align_right
            sheet.row_dimensions[current_row].height = 21
            current_row += 1
            
        # --- SECTION II: VELOCITY TRAJECTORY ---
        current_row += 2
        sheet.cell(row=current_row, column=2, value="II. Recent 14-Day Velocity & Output Analysis").font = font_section
        current_row += 1
        
        headers_growth = list(df_growth.columns)
        for col_idx, header in enumerate(headers_growth, start=2):
            cell = sheet.cell(row=current_row, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_indigo
            cell.alignment = align_center
            cell.border = border_saffron_top
        sheet.row_dimensions[current_row].height = 26
        current_row += 1
        
        for row_idx, row_data in enumerate(df_growth.values):
            row_fill = fill_white if row_idx % 2 == 0 else fill_zebra
            for c_idx, value in enumerate(row_data, start=2):
                cell = sheet.cell(row=current_row, column=c_idx, value=value)
                cell.font = font_data
                cell.border = thin_border
                cell.fill = row_fill
                
                if c_idx in [2, 3]:
                    cell.alignment = align_center
                elif isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = align_right
                else:
                    cell.alignment = align_center
            sheet.row_dimensions[current_row].height = 20
            current_row += 1
            
        # --- AUTO-FIT LOGIC ---
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            if col_letter == 'A':
                sheet.column_dimensions[col_letter].width = 3
                continue
            for cell in col:
                if cell.row > 3 and cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 16)
            
    print(f"\n✨ Generation Successful! File written cleanly to: '{filename}'")

if __name__ == "__main__":
    build_political_dashboard()